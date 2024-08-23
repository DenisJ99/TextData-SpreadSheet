import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def adjust_column_widths(sheet):
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

def extract_data(text):
    data = {}
    process_names = {}
    event_counts = { 'THRECEIVE': {}, 'THCONDVAR': {}, 'THREPLY': {}, 'THSEM': {}, 'THMUTEX': {}, 'THNANOSLEEP': {} }
    
    current_pid = current_tid = current_name = None
    
    for line in text.split('\n'):
        pid_match = re.search(r'pid:(\d+)', line)
        if pid_match:
            current_pid = pid_match.group(1)
            if current_pid not in data:
                data[current_pid] = {}
            current_tid = current_name = None
        
        if current_pid:
            tid_match = re.search(r'tid:(\d+)', line)
            if tid_match:
                current_tid = tid_match.group(1)
                if current_tid not in data[current_pid]:
                    data[current_pid][current_tid] = "Unnamed Thread"
            
            name_match = re.search(r'name:(.+)', line)
            if name_match:
                current_name = name_match.group(1).strip()
                if current_tid is None:
                    if current_pid not in process_names:
                        process_names[current_pid] = current_name
                else:
                    data[current_pid][current_tid] = current_name

            for event in event_counts:
                if event in line and current_tid:
                    event_counts[event].setdefault(current_pid, {}).setdefault(current_tid, 0)
                    event_counts[event][current_pid][current_tid] += 1
    
    return data, process_names, event_counts

def set_cell_style(cell, font=None, alignment=None, fill=None, border=None):
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = fill
    if border:
        cell.border = border

def write_to_xlsx(data, output_file, process_names, event_counts):
    workbook = Workbook()

    for pid, threads in data.items():
        sheet = workbook.create_sheet(title=f"PID_{pid}")
        process_name = process_names.get(pid, "Unknown Process")
        total_counts = {event: 0 for event in event_counts}
        
        # Define styles
        header_font = Font(bold=True)
        center_aligned_text = Alignment(horizontal='center')
        fill_green = PatternFill(start_color='96d2b8', end_color='96d2b8', fill_type='solid')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Write headers for the first row
        headers_row1 = ["Process Name", "Process ID", "", "Running Time (msec)", "CPU Usage %",
                        "THRECEIVE", "THCONDVAR", "THREPLY", "THSEM", "THMUTEX", "THNANOSLEEP"]
        for col, header in enumerate(headers_row1, start=1):
            cell = sheet.cell(row=1, column=col)
            cell.value = header
            set_cell_style(cell, font=header_font, alignment=center_aligned_text, fill=fill_green, border=thin_border)
        
        # Write headers for the third row
        headers_row3 = ["Thread Name", "Thread ID", "Thread Owner", "Running Time (msec)", "CPU Usage %",
                        "THRECEIVE", "THCONDVAR", "THREPLY", "THSEM", "THMUTEX", "THNANOSLEEP"]
        for col, header in enumerate(headers_row3, start=1):
            cell = sheet.cell(row=3, column=col)
            cell.value = header
            set_cell_style(cell, font=header_font, alignment=center_aligned_text, fill=fill_green, border=thin_border)
        
        # Add filter
        sheet.auto_filter.ref = f"A3:{get_column_letter(len(headers_row3))}{len(threads) + 3}"
        
        row_offset = 4
        # Write the process name and PID
        cell = sheet.cell(row=2, column=1)
        cell.value = process_name
        cell = sheet.cell(row=2, column=2)
        cell.value = pid
        
        for tid, thread_name in threads.items():
            cell = sheet.cell(row=row_offset, column=1)
            cell.value = thread_name
            
            cell = sheet.cell(row=row_offset, column=2)
            cell.value = int(tid)  # Ensure Thread ID is treated as a number for correct sorting
            
            for col, event in enumerate(headers_row3[5:], start=6):
                count = event_counts[event].get(pid, {}).get(tid, 0)
                total_counts[event] += count
                cell = sheet.cell(row=row_offset, column=col)
                cell.value = count if count != 0 else ""
            
            row_offset += 1
        
        # Write total counts for each event in the second row
        for col, event in enumerate(headers_row3[5:], start=6):
            cell = sheet.cell(row=2, column=col)
            cell.value = total_counts[event] if total_counts[event] != 0 else ""  
        
        adjust_column_widths(sheet)
    
    workbook.remove(workbook['Sheet'])  # Remove the default sheet created by Workbook
    workbook.save(output_file)

def main():
    input_file = input("Please enter the name of the input text file: ")
    output_file = input("Please enter the name of the output XLSX file: ")

    with open(input_file, 'r') as file:
        text = file.read()

    extracted_data, process_names, event_counts = extract_data(text)
    write_to_xlsx(extracted_data, output_file, process_names, event_counts)
    print(f"Data has been written to {output_file}.")

if __name__ == "__main__":
    main()
